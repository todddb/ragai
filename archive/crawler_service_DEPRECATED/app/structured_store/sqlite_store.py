from __future__ import annotations

import io
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict


@dataclass
class XlsxIngestResult:
    sheet_count: int
    cell_count: int
    truncated: bool


class SQLiteStructuredStore:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path

    def ensure_schema(self, conn: sqlite3.Connection) -> None:
        schema_path = Path(__file__).with_name("schema.sql")
        conn.executescript(schema_path.read_text(encoding="utf-8"))

    def ingest_xlsx(
        self,
        doc_id: str,
        source_url: str,
        content_bytes: bytes,
        max_cells: int,
        batch_size: int,
    ) -> XlsxIngestResult:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=io.BytesIO(content_bytes), data_only=True)
        ingested_at = datetime.utcnow().isoformat() + "Z"
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        cell_count = 0
        truncated = False
        with sqlite3.connect(self.db_path) as conn:
            self.ensure_schema(conn)
            conn.execute(
                "INSERT OR REPLACE INTO documents (doc_id, source_url, ingested_at) VALUES (?, ?, ?)",
                (doc_id, source_url, ingested_at),
            )
            conn.execute(
                """
                INSERT OR REPLACE INTO xlsx_workbooks
                (doc_id, source_url, sheet_count, truncated, ingested_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (doc_id, source_url, len(workbook.worksheets), 0, ingested_at),
            )
            for sheet_index, sheet in enumerate(workbook.worksheets):
                conn.execute(
                    """
                    INSERT OR REPLACE INTO xlsx_sheets (doc_id, sheet_name, sheet_index)
                    VALUES (?, ?, ?)
                    """,
                    (doc_id, sheet.title, sheet_index),
                )
                batch = []
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    for col_index, value in enumerate(row, start=1):
                        if value is None:
                            continue
                        if isinstance(value, str) and not value.strip():
                            continue
                        batch.append((doc_id, sheet.title, row_index, col_index, str(value)))
                        cell_count += 1
                        if len(batch) >= batch_size:
                            conn.executemany(
                                """
                                INSERT INTO xlsx_cells (doc_id, sheet_name, row, column, value)
                                VALUES (?, ?, ?, ?, ?)
                                """,
                                batch,
                            )
                            batch.clear()
                        if cell_count >= max_cells:
                            truncated = True
                            break
                    if truncated:
                        break
                if batch:
                    conn.executemany(
                        """
                        INSERT INTO xlsx_cells (doc_id, sheet_name, row, column, value)
                        VALUES (?, ?, ?, ?, ?)
                        """,
                        batch,
                    )
                if truncated:
                    break
            if truncated:
                conn.execute(
                    "UPDATE xlsx_workbooks SET truncated = 1 WHERE doc_id = ?",
                    (doc_id,),
                )
            conn.commit()
        return XlsxIngestResult(
            sheet_count=len(workbook.worksheets),
            cell_count=cell_count,
            truncated=truncated,
        )

    def ingest_xlsx_to_meta(
        self,
        doc_id: str,
        source_url: str,
        content_bytes: bytes,
        max_cells: int,
        batch_size: int,
    ) -> Dict[str, int | bool]:
        result = self.ingest_xlsx(doc_id, source_url, content_bytes, max_cells, batch_size)
        return {
            "sheet_count": result.sheet_count,
            "cell_count": result.cell_count,
            "truncated": result.truncated,
        }
