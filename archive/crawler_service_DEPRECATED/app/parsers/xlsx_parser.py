from __future__ import annotations

import io
from typing import List

from openpyxl import load_workbook

from app.parsers.types import ParsedDocument


def parse_xlsx(content_bytes: bytes) -> ParsedDocument:
    workbook = load_workbook(filename=io.BytesIO(content_bytes), data_only=True)
    markdown_lines: List[str] = []
    text_parts: List[str] = []
    for sheet in workbook.worksheets:
        markdown_lines.append(f"# {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if not any(value.strip() for value in values):
                continue
            markdown_lines.append(" | ".join(values))
            text_parts.extend(value for value in values if value.strip())
    markdown = "\n".join(markdown_lines).strip()
    text = " ".join(text_parts)
    return ParsedDocument(
        title="",
        markdown=markdown,
        text_for_chunking=text,
        meta={"sheet_count": len(workbook.worksheets)},
    )
